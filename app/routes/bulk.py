"""Bulk import/export routes — CSV/Excel upload with per-branch scoping."""
from __future__ import annotations

import io
import uuid
from collections import Counter
from datetime import date, datetime, timezone

import pandas as pd
import structlog
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.database import BulkImport, HsnCode, Prediction, User, UserRole, get_db
from app.routes.auth import require_role
from app.services.audit import EventType, log_event
from app.services.confidence import score_result
from app.services.db_matcher import match_query
from app.services.matcher import get_matcher
from app.routes.predict import _build_gst_fields

router = APIRouter(tags=["bulk"])
log = structlog.get_logger()

MAX_FILE_SIZE = 5 * 1024 * 1024   # 5 MB
MAX_ROWS = 1000


@router.post("/predict/bulk/upload")
async def bulk_upload(
    file: UploadFile = File(...),
    current_user: User = Depends(
        require_role(
            UserRole.BRANCH_USER,
            UserRole.BRANCH_MANAGER,
            UserRole.REGIONAL_ADMIN,
            UserRole.HQ_ADMIN,
            UserRole.AUDITOR,
        )
    ),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()

    # 1. Extension check
    filename = file.filename or "upload"
    fname_lower = filename.lower()
    if not (fname_lower.endswith(".csv") or fname_lower.endswith(".xlsx")):
        raise HTTPException(status_code=422, detail="Only .csv and .xlsx files are supported")

    # 2. Size check
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File exceeds 5 MB limit")

    # 3. Parse
    if fname_lower.endswith(".xlsx"):
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
    else:
        df = pd.read_csv(io.BytesIO(content), encoding="utf-8", errors="replace")

    df.columns = [c.strip().lower() for c in df.columns]

    # 4. Column presence check
    if "product_description" not in df.columns:
        raise HTTPException(status_code=422, detail="Missing required column: product_description")

    # 5. Row count check
    if len(df) > MAX_ROWS:
        raise HTTPException(status_code=422, detail=f"File contains {len(df)} rows; maximum is {MAX_ROWS}")

    import_row = BulkImport(
        branch_id=current_user.branch_id,
        user_id=current_user.id,
        filename=filename,
        row_count=len(df),
        status="processing",
    )
    db.add(import_row)
    await db.flush()

    results = []
    for idx, row in df.iterrows():
        text = str(row.get("product_description", "")).strip()
        if not text or text.lower() in ("nan", "none"):
            continue

        matches = await match_query(text, db, top_k=5)
        if not matches:
            matcher = get_matcher()
            matches = matcher.match(text, top_k=5)
        if not matches:
            continue

        top = matches[0]
        confidence, label = score_result(top["score"])
        needs_review = confidence < 0.7
        gst_fields = await _build_gst_fields(top["hsn_code"], db)

        pred = Prediction(
            request_id=str(uuid.uuid4()),
            input_text=text,
            predicted_hsn=top["hsn_code"],
            confidence=confidence,
            needs_review=needs_review,
            api_key_hash=None,
            branch_id=current_user.branch_id,
        )
        db.add(pred)
        await db.flush()

        results.append({
            "row_index": int(idx),
            "product_description": text,
            "hsn_code": top["hsn_code"],
            "confidence": confidence,
            "confidence_label": label,
            "gst_rate": gst_fields["gst_rate"],
            "gst_effective_from": gst_fields["gst_effective_from"],
            "needs_review": needs_review,
        })

    import_row.status = "completed"
    import_row.completed_at = datetime.now(timezone.utc)
    import_row.row_count = len(df)

    await log_event(
        session=db,
        event_type=EventType.BULK_IMPORT,
        actor_user_id=current_user.id,
        actor_role=current_user.role,
        branch_id=current_user.branch_id,
        entity_type="bulk_import",
        entity_id=str(import_row.id),
        new_value={
            "filename": filename,
            "row_count": len(df),
            "branch_id": str(current_user.branch_id),
        },
    )
    await db.commit()

    return {
        "import_id": str(import_row.id),
        "row_count": len(df),
        "results": results,
    }


@router.get("/predict/bulk/{import_id}/export")
async def bulk_export(
    import_id: str,
    current_user: User = Depends(
        require_role(
            UserRole.BRANCH_USER,
            UserRole.BRANCH_MANAGER,
            UserRole.REGIONAL_ADMIN,
            UserRole.HQ_ADMIN,
            UserRole.AUDITOR,
        )
    ),
    db: AsyncSession = Depends(get_db),
):
    import_row = (
        await db.execute(select(BulkImport).where(BulkImport.id == import_id))
    ).scalars().first()

    if not import_row:
        raise HTTPException(status_code=404, detail="Bulk import not found")

    if (
        current_user.role != UserRole.HQ_ADMIN.value
        and import_row.branch_id != current_user.branch_id
    ):
        raise HTTPException(status_code=403, detail="Not allowed")

    predictions = (
        await db.execute(
            select(Prediction).where(
                Prediction.branch_id == import_row.branch_id,
                Prediction.created_at >= import_row.created_at,
            )
        )
    ).scalars().all()

    # Fetch branch name for filename
    branch_name = str(import_row.branch_id)[:8]

    wb = Workbook()

    # Sheet 1: Classifications
    ws1 = wb.active
    ws1.title = "Classifications"
    headers = [
        "Product Description", "HSN Code", "Confidence",
        "Confidence Label", "GST Rate (%)", "Effective From", "Needs Review",
    ]
    ws1.append(headers)
    fill = PatternFill("solid", fgColor="1a56db")
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for pred in predictions:
        label = "HIGH" if pred.confidence >= 0.8 else ("MEDIUM" if pred.confidence >= 0.55 else "LOW")
        gst_row = (
            await db.execute(
                select(HsnCode.gst_rate_numeric, HsnCode.gst_effective_from)
                .where(HsnCode.hsn_code == pred.predicted_hsn)
            )
        ).first()
        rate = float(gst_row.gst_rate_numeric) if gst_row and gst_row.gst_rate_numeric is not None else None
        eff_from = str(gst_row.gst_effective_from) if gst_row and gst_row.gst_effective_from else None
        ws1.append([
            pred.input_text,
            pred.predicted_hsn,
            round(pred.confidence, 3),
            label,
            rate,
            eff_from,
            pred.needs_review,
        ])

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = (
            max(len(str(c.value or "")) for c in col) + 4
        )

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    high = sum(1 for p in predictions if p.confidence >= 0.7)
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Rows", len(predictions)])
    ws2.append(["High Confidence (≥0.7)", high])
    ws2.append(["Low Confidence (<0.7)", len(predictions) - high])

    gst_rates: list[float | None] = []
    for pred in predictions:
        gst_row = (
            await db.execute(
                select(HsnCode.gst_rate_numeric)
                .where(HsnCode.hsn_code == pred.predicted_hsn)
            )
        ).first()
        gst_rates.append(
            float(gst_row.gst_rate_numeric) if gst_row and gst_row.gst_rate_numeric is not None else None
        )

    for rate, count in sorted(Counter(r for r in gst_rates if r is not None).items()):
        ws2.append([f"GST {int(rate)}%", count])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"hsn_classification_{branch_name}_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
