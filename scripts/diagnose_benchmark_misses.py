#!/usr/bin/env python3
"""Diagnose why benchmark rows fail _is_detected."""
from __future__ import annotations

import asyncio
import json
import os
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)


def _load_env() -> None:
    p = ROOT / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _why(result: dict) -> str:
    hsn = (result.get("hsn_code") or "").strip()
    if not hsn or hsn in ("UNKNOWN", "UNCLASSIFIED", "99999999"):
        return "invalid_hsn"
    digits = "".join(c for c in hsn if c.isdigit())
    if len(digits) not in (4, 6, 8):
        return "bad_hsn_len"
    if result.get("gst_rate") is None:
        return "gst_null"
    if int(result.get("confidence_score") or result.get("confidence") or 0) < 70:
        return "low_conf"
    if result.get("needs_manual_review") or result.get("review_required"):
        if result.get("rate_conflict"):
            return "rate_conflict"
        return "review_flag"
    return "ok"


async def run(excel: Path, limit: int | None = 91) -> None:
    _load_env()
    os.environ["FAISS_DISABLED"] = "1"
    from openpyxl import load_workbook

    from app.models.database import async_session, init_db
    from app.services.gst_classifier import classify

    names: list[str] = []
    wb = load_workbook(excel, read_only=True)
    for i, row in enumerate(wb.active.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row and row[0]:
            names.append(str(row[0]).strip())
    wb.close()
    if limit:
        names = names[:limit]

    await init_db()
    reasons: Counter[str] = Counter()
    rows_out: list[dict] = []

    for desc in names:
        async with async_session() as db:
            out = await classify(db, desc, bypass_cache=True)
        reason = _why(out)
        reasons[reason] += 1
        rows_out.append({
            "description": desc,
            "reason": reason,
            "hsn_code": out.get("hsn_code"),
            "gst_rate": out.get("gst_rate"),
            "confidence": out.get("confidence"),
            "review_required": out.get("review_required"),
            "rate_conflict": out.get("rate_conflict"),
            "matched_layer": out.get("matched_layer"),
            "tier_used": out.get("tier_used"),
        })

    print(json.dumps({"total": len(names), "reasons": dict(reasons)}, indent=2))
    missed = [r for r in rows_out if r["reason"] != "ok"]
    print(f"\nMissed: {len(missed)}/{len(names)}")
    for r in missed[:15]:
        print(
            f"  {r['description'][:45]:45} | {r['reason']:14} | "
            f"{r['hsn_code']} gst={r['gst_rate']} conf={r['confidence']} "
            f"rc={r['rate_conflict']} layer={r['matched_layer']}"
        )
    out = ROOT / "scripts" / "kerala_miss_diagnosis.json"
    out.write_text(json.dumps({"reasons": dict(reasons), "rows": rows_out}, indent=2), encoding="utf-8")
    print(f"\nWrote {out}")


if __name__ == "__main__":
    excel = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "kerala_invoice_benchmark.xlsx"
    lim = int(sys.argv[2]) if len(sys.argv) > 2 else None
    asyncio.run(run(excel, lim))
